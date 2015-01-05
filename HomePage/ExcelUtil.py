# -*- coding:utf-8 -*-
from __future__ import division

import xlrd,xlwt
import openpyxl
import os
#万恶的ExcelWriter，妹的封装好了不早说，封装了很强大的excel写的功能
from openpyxl.writer.excel import ExcelWriter
#一个eggache的数字转为列字母的方法
from openpyxl.cell import get_column_letter
# 样式
from openpyxl.styles import Style, PatternFill, Border, Side, Alignment, Font
import copy


__author__ = 'kiven'

class Excel_2007_Engine(object):

    def __init__(self,excel_name):
        # excel 名称
        self.excel_name = 'Download/' + excel_name
        self.excel_suffix = u'测试报告.xlsx'
        # excel book对象
        if os.path.exists(self.excel_name):
            self.work_book = openpyxl.load_workbook()
        else:
            self.work_book = openpyxl.Workbook()
        # 预定义sheets名称
        self.sheet_names = [u'接口测试设计明细',u'接口测试结果明细',u'block问题统计',u'问题描述']


    # 创建excel
    def create_excel(self,title,interface_info,all_case_num,all_pass_case_num,all_check_num):
        # 新建一个ExcelWriter
        excel_work = ExcelWriter(workbook = self.work_book)
        # 设置文件输出路径与名称
        dest_filename = self.excel_name + self.excel_suffix
        # 创建sheet
        for sheet_name in self.sheet_names:
            ws = self.work_book.create_sheet()
            ws.title = sheet_name

        # ==============================================================================================================
        # 根据sheet名称获取该sheet对象
        # 第一个sheet
        ws = self.work_book.get_sheet_by_name(self.sheet_names[0])

        '''
        合并单元格
        源码中该方法的定义
        def merge_cells(self, range_string=None, start_row=None, start_column=None, end_row=None, end_column=None)
        所以可以直接使用后面四个参数
        '''
        ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=9)
        c_a1 = ws['A1']
        c_a3 = ws['A3']
        c_b3 = ws['B3']
        c_c3 = ws['C3']
        c_d3 = ws['D3']
        c_e3 = ws['E3']
        c_f3 = ws['F3']
        c_g3 = ws['G3']
        c_h3 = ws['H3']
        c_i3 = ws['I3']

        # 接口测试设计明细 sheet页的通用样式
        general_style = Style(
            font=Font(
                name=u'微软雅黑',
                size=10
            ),
            fill=PatternFill(
                start_color='DCE6F1',
                end_color='DCE6F1',
                fill_type='solid'
            ),
            border=Border(
                left=Side(color='000000',border_style='thin'),
                right=Side(color='000000',border_style='thin'),
                top=Side(color='000000',border_style='thin'),
                bottom=Side(color='000000',border_style='thin'),
                diagonal=Side(color='000000',border_style='thin'),
            )
        )
        c_a3.value = u'涉及模块'
        c_b3.value = u'接口数量'
        c_c3.value = u'接口名称'
        c_d3.value = u'测试Owner'
        c_e3.value = u'契约文档提供'
        c_f3.value = u'契约文档确认'
        c_g3.value = u'SOA接口是否提供'
        c_h3.value = u'预计用例数量'
        c_i3.value = u'完成用例数量'

        # 设置成通用样式
        c_a3.style = c_b3.style = c_c3.style = c_d3.style = c_e3.style = c_f3.style = c_g3.style = c_h3.style = c_i3.style = general_style

        # 单元格值
        c_a1.value = title + u'接口测试明细'
        # 单元格样式
        c_a1.style = Style(
            # 字体
            font=Font(
                name=u'微软雅黑',
                size=22,
            ),
            # 对齐方式
            alignment = Alignment(
                # 水平对齐   居中
                horizontal = 'center'
            ),
            # 填充
            fill=PatternFill(
                start_color='FFC000',
                end_color='FFC000',
                fill_type='solid'
            ),
            border=Border(
                left=Side(color='000000',border_style='thin'),
                right=Side(color='000000',border_style='thin'),
                top=Side(color='000000',border_style='thin'),
                bottom=Side(color='000000',border_style='thin'),
                diagonal=Side(color='000000',border_style='thin'),
            )
        )
        # ==============================================================================================================



        # ==============================================================================================================
        # 第二个sheet
        ws = self.work_book.get_sheet_by_name(self.sheet_names[1])
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)
        ws.merge_cells(start_row=2, start_column=1, end_row=3, end_column=1)
        ws.merge_cells(start_row=2, start_column=2, end_row=2, end_column=6)
        c_a1 = ws['A1']
        c_a2 = ws['A2']
        c_b2 = ws['B2']
        c_b3 = ws['B3']
        c_c3 = ws['C3']
        c_d3 = ws['D3']
        c_e3 = ws['E3']
        c_f3 = ws['F3']

        c_a1.value = title + u'接口测试数据统计'
        c_a2.value = title
        c_b2.value = u'验收结果'
        c_b3.value = u'测试用例'
        c_c3.value = u'通过用例'
        c_d3.value = u'失败用例'
        c_e3.value = u'检查点数量'
        c_f3.value = u'通过率'

        c_a1.style = Style(
            font=Font(
                name=u'微软雅黑',
                size=11,
                color='FFFFFF',
                bold='bold'
            ),
            # 对齐方式
            alignment = Alignment(
                # 水平对齐   居中
                horizontal = 'center'
            ),
            # 填充
            fill=PatternFill(
                start_color='0070C0',
                end_color='0070C0',
                fill_type='solid'
            ),
        )
        c_a2.style = Style(
            font=Font(
                name=u'微软雅黑',
                size=11
            ),
            # 对齐方式
            alignment = Alignment(
                # 水平对齐   居中
                horizontal = 'center',
                # 垂直居中
                vertical = 'center',
            ),
            # 填充
            fill=PatternFill(
                start_color='00B050',
                end_color='00B050',
                fill_type='solid'
            ),
            border=Border(
                left=Side(color='000000',border_style='thin'),
                right=Side(color='000000',border_style='thin'),
                top=Side(color='000000',border_style='thin'),
                bottom=Side(color='000000',border_style='thin'),
                diagonal=Side(color='000000',border_style='thin'),
            )
        )
        c_b2.style = Style(
            font=Font(
                name=u'微软雅黑',
                size=11
            ),
            # 对齐方式
            alignment = Alignment(
                # 水平对齐   居中
                horizontal = 'center',
            ),
            # 填充
            fill=PatternFill(
                start_color='FCD5B4',
                end_color='FCD5B4',
                fill_type='solid'
            ),
            border=Border(
                left=Side(color='000000',border_style='thin'),
                right=Side(color='000000',border_style='thin'),
                top=Side(color='000000',border_style='thin'),
                bottom=Side(color='000000',border_style='thin'),
                diagonal=Side(color='000000',border_style='thin'),
            )
        )
        c_b3.style = Style(
            font=Font(
                name=u'微软雅黑',
                size=11
            ),
            # 对齐方式
            alignment = Alignment(
                # 水平对齐   居中
                horizontal = 'center',
            ),
            border=Border(
                left=Side(color='000000',border_style='thin'),
                right=Side(color='000000',border_style='thin'),
                top=Side(color='000000',border_style='thin'),
                bottom=Side(color='000000',border_style='thin'),
                diagonal=Side(color='000000',border_style='thin'),
            )
        )
        c_e3.style = Style(
            font=Font(
                name=u'微软雅黑',
                size=11
            ),
            # 对齐方式
            alignment = Alignment(
                # 水平对齐   居中
                horizontal = 'center',
            ),
            border=Border(
                left=Side(color='000000',border_style='thin'),
                right=Side(color='000000',border_style='thin'),
                top=Side(color='000000',border_style='thin'),
                bottom=Side(color='000000',border_style='thin'),
                diagonal=Side(color='000000',border_style='thin'),
            )
        )
        c_c3.style = Style(
            font=Font(
                name=u'微软雅黑',
                size=11
            ),
            # 对齐方式
            alignment = Alignment(
                # 水平对齐   居中
                horizontal = 'center',
            ),
            # 填充
            fill=PatternFill(
                start_color='92D050',
                end_color='92D050',
                fill_type='solid'
            ),
            border=Border(
                left=Side(color='000000',border_style='thin'),
                right=Side(color='000000',border_style='thin'),
                top=Side(color='000000',border_style='thin'),
                bottom=Side(color='000000',border_style='thin'),
                diagonal=Side(color='000000',border_style='thin'),
            )
        )
        c_d3.style = Style(
            font=Font(
                name=u'微软雅黑',
                size=11
            ),
            # 对齐方式
            alignment = Alignment(
                # 水平对齐   居中
                horizontal = 'center',
            ),
            # 填充
            fill=PatternFill(
                start_color='FF0000',
                end_color='FF0000',
                fill_type='solid'
            ),
            border=Border(
                left=Side(color='000000',border_style='thin'),
                right=Side(color='000000',border_style='thin'),
                top=Side(color='000000',border_style='thin'),
                bottom=Side(color='000000',border_style='thin'),
                diagonal=Side(color='000000',border_style='thin'),
            )
        )
        c_f3.style = Style(
            font=Font(
                name=u'微软雅黑',
                size=11
            ),
            # 对齐方式
            alignment = Alignment(
                # 水平对齐   居中
                horizontal = 'center',
            ),
            # 填充
            fill=PatternFill(
                start_color='DAEEF3',
                end_color='DAEEF3',
                fill_type='solid'
            ),
            border=Border(
                left=Side(color='000000',border_style='thin'),
                right=Side(color='000000',border_style='thin'),
                top=Side(color='000000',border_style='thin'),
                bottom=Side(color='000000',border_style='thin'),
                diagonal=Side(color='000000',border_style='thin'),
            )
        )
        # ==============================================================================================================



        # ==============================================================================================================
        # 第三个sheet
        ws = self.work_book.get_sheet_by_name(self.sheet_names[2])
        ws.merge_cells(start_row=2, start_column=1, end_row=3, end_column=1)
        ws.merge_cells(start_row=2, start_column=2, end_row=3, end_column=2)
        c_a1 = ws['A1']
        c_b1 = ws['B1']
        c_a1.value = u'涉及接口'
        c_b1.value = u'问题说明'
        c_a1.style = c_b1.style = Style(
            font=Font(
                name=u'微软雅黑',
                size=11
            ),
            # 对齐方式
            alignment = Alignment(
                # 水平对齐   居中
                horizontal = 'center',
            ),
            # 填充
            fill=PatternFill(
                start_color='DCE6F1',
                end_color='DCE6F1',
                fill_type='solid'
            ),
            border=Border(
                left=Side(color='000000',border_style='thin'),
                right=Side(color='000000',border_style='thin'),
                top=Side(color='000000',border_style='thin'),
                bottom=Side(color='000000',border_style='thin'),
                diagonal=Side(color='000000',border_style='thin'),
            )
        )
        # ==============================================================================================================




        # ==============================================================================================================
        # 第四个sheet
        ws = self.work_book.get_sheet_by_name(self.sheet_names[3])
        c_a1 = ws['A1']
        c_b1 = ws['B1']
        c_c1 = ws['C1']
        c_d1 = ws['D1']
        c_e1 = ws['E1']
        c_a6 = ws['A6']
        c_a7 = ws['A7']
        c_b7 = ws['B7']
        c_c7 = ws['C7']
        c_d7 = ws['D7']
        c_e7 = ws['E7']
        c_a10 = ws['A10']

        c_a1.value = u'模块'
        c_b1.value = u'概述'
        c_c1.value = u'问题归属'
        c_d1.value = u'备注'
        c_e1.value = u'问题影响'
        c_a6.value = u'今日已修复问题'
        c_a7.value = u'模块'
        c_b7.value = u'概述'
        c_c7.value = u'问题归属'
        c_d7.value = u'备注'
        c_a10.value = u'无需解决问题'
        c_a1.style = c_b1.style = c_c1.style = c_d1.style = c_e1.style = c_a7.style = c_b7.style = c_c7.style = c_d7.style = c_e7.style = Style(
            font=Font(
                name=u'微软雅黑',
                size=10
            ),
            # 对齐方式
            alignment = Alignment(
                # 水平对齐   居中
                horizontal = 'center',
            ),
            # 填充
            fill=PatternFill(
                start_color='DDD9C4',
                end_color='DDD9C4',
                fill_type='solid'
            ),
            border=Border(
                left=Side(color='000000',border_style='thin'),
                right=Side(color='000000',border_style='thin'),
                top=Side(color='000000',border_style='thin'),
                bottom=Side(color='000000',border_style='thin'),
                diagonal=Side(color='000000',border_style='thin'),
            )
        )
        c_a6.style =  c_a10.style = Style(
            font=Font(
                name=u'微软雅黑',
                size=10
            ),
            # 对齐方式
            alignment = Alignment(
                # 水平对齐   居中
                horizontal = 'center',
            ),
            # 填充
            fill=PatternFill(
                start_color='FFFF00',
                end_color='FFFF00',
                fill_type='solid'
            ),
        )
        # ==============================================================================================================
        # 传入一个对象
        #
        # interface_info = [
        #     {
        #         'name':u"这是一个接口名称1",
        #         'owner':u'沈佳龙',
        #         'contract_is_provide':u'是',
        #         'is_confirm':u'是',
        #         'soa_is_provide':u'是',
        #         'case_num':10
        #     },
        #     {
        #         'name':u"这是一个接口名称2",
        #         'owner':u'张三',
        #         'contract_is_provide':u'是',
        #         'is_confirm':u'是',
        #         'soa_is_provide':u'是',
        #         'case_num':20
        #     },
        #     {
        #         'name':u"这是一个接口名称3",
        #         'owner':u'李四',
        #         'contract_is_provide':u'是',
        #         'is_confirm':u'是',
        #         'soa_is_provide':u'是',
        #         'case_num':20
        #     },
        # ]

        # 调用 self.insert_interface_info(case_info)
        self.insert_interface_info(title,interface_info,all_case_num,all_pass_case_num,all_check_num)
        excel_work.save(filename=dest_filename)

    # 插入接口数据
    def insert_interface_info(self,title,interface_info,all_case_num,all_pass_case_num,all_check_num):
        tmp = copy.deepcopy(interface_info) # 对象深拷贝
        # 接口的数量,首先渲染整体样式
        interface_num = len(interface_info)
        # ==============================================================================================================
        work_sheet = self.work_book.get_sheet_by_name(self.sheet_names[0])
        work_sheet.merge_cells(start_row=4, start_column=1, end_row=interface_num+3, end_column=1)
        work_sheet.merge_cells(start_row=4, start_column=2, end_row=interface_num+3, end_column=2)
        # 涉及模块
        work_sheet.cell(row=4,column=1).value = title
        # 接口数量
        work_sheet.cell(row=4,column=2).value = interface_num
        # 其他信息
        for row in range(4,interface_num+4):
            for i in interface_info:
                work_sheet.cell(row=row,column=3).value = i['name']
                work_sheet.cell(row=row,column=4).value = i['owner']
                work_sheet.cell(row=row,column=5).value = i['contract_is_provide']
                work_sheet.cell(row=row,column=6).value = i['is_confirm']
                work_sheet.cell(row=row,column=7).value = i['soa_is_provide']
                work_sheet.cell(row=row,column=8).value = i['case_num']
                work_sheet.cell(row=row,column=9).value = i['case_num']
                # 填完一个后将它从列表中remove掉
                interface_info.remove(i)
                # 直接跳出内循环
                break

        for row in range(4,interface_num+4):
            for col in range(1,10):
                work_sheet.cell(row=row,column=col).style = Style(
                    font=Font(
                        name=u'微软雅黑',
                        size=10
                    ),
                    # 对齐方式
                    alignment = Alignment(
                        # 水平对齐   居中
                        horizontal = 'center',
                    ),
                    # 填充
                    fill=PatternFill(
                        start_color='DCE6F1',
                        end_color='DCE6F1',
                        fill_type='solid'
                    ),
                    border=Border(
                        left=Side(color='000000',border_style='thin'),
                        right=Side(color='000000',border_style='thin'),
                        top=Side(color='000000',border_style='thin'),
                        bottom=Side(color='000000',border_style='thin'),
                        diagonal=Side(color='000000',border_style='thin'),
                    )
                )
        # ==============================================================================================================


        # ==============================================================================================================
        work_sheet = self.work_book.get_sheet_by_name(self.sheet_names[1])
        # 接口数据
        for row in range(4,interface_num+4):
            for i in tmp:# tmp是interface_info对象的深拷贝
                work_sheet.cell(row=row,column=1).value = i['name']
                work_sheet.cell(row=row,column=2).value = i['case_num']
                work_sheet.cell(row=row,column=3).value = i['pass_case_num']
                work_sheet.cell(row=row,column=4).value = i['unpass_case_num']
                work_sheet.cell(row=row,column=5).value = i['check_num']
                work_sheet.cell(row=row,column=6).value = i['pass_rate'] + '%'
                # 填完一个后将它从列表中remove掉
                tmp.remove(i)
                # 直接跳出内循环
                break
        # 总计一栏
        row_num = interface_num+4
        work_sheet.cell(row=row_num,column=1).value = u'总计' # 总计单元格EBF1DE
        work_sheet.cell(row=row_num,column=1).style = Style(
            font=Font(
                name=u'微软雅黑',
                size=10
            ),
            fill=PatternFill(
                start_color='EBF1DE',
                end_color='EBF1DE',
                fill_type='solid'
            ),
            border=Border(
                left=Side(color='000000',border_style='thin'),
                right=Side(color='000000',border_style='thin'),
                top=Side(color='000000',border_style='thin'),
                bottom=Side(color='000000',border_style='thin'),
                diagonal=Side(color='000000',border_style='thin'),
            )
        )

        work_sheet.cell(row=row_num,column=2).value = all_case_num
        work_sheet.cell(row=row_num,column=2).style = Style(
            font=Font(
                name=u'微软雅黑',
                size=10
            ),
            border=Border(
                left=Side(color='000000',border_style='thin'),
                right=Side(color='000000',border_style='thin'),
                top=Side(color='000000',border_style='thin'),
                bottom=Side(color='000000',border_style='thin'),
                diagonal=Side(color='000000',border_style='thin'),
            )
        )

        work_sheet.cell(row=row_num,column=3).value = all_pass_case_num
        work_sheet.cell(row=row_num,column=3).style = Style(
            font=Font(
                name=u'微软雅黑',
                size=10
            ),
            fill=PatternFill(
                start_color='92D050',
                end_color='92D050',
                fill_type='solid'
            ),
            border=Border(
                left=Side(color='000000',border_style='thin'),
                right=Side(color='000000',border_style='thin'),
                top=Side(color='000000',border_style='thin'),
                bottom=Side(color='000000',border_style='thin'),
                diagonal=Side(color='000000',border_style='thin'),
            )
        )

        work_sheet.cell(row=row_num,column=4).value = all_case_num-all_pass_case_num
        work_sheet.cell(row=row_num,column=4).style = Style(
            font=Font(
                name=u'微软雅黑',
                size=10
            ),
            fill=PatternFill(
                start_color='FF0000',
                end_color='FF0000',
                fill_type='solid'
            ),
            border=Border(
                left=Side(color='000000',border_style='thin'),
                right=Side(color='000000',border_style='thin'),
                top=Side(color='000000',border_style='thin'),
                bottom=Side(color='000000',border_style='thin'),
                diagonal=Side(color='000000',border_style='thin'),
            )
        )


        work_sheet.cell(row=row_num,column=5).value = all_check_num
        work_sheet.cell(row=row_num,column=5).style = Style(
            font=Font(
                name=u'微软雅黑',
                size=10
            ),
            border=Border(
                left=Side(color='000000',border_style='thin'),
                right=Side(color='000000',border_style='thin'),
                top=Side(color='000000',border_style='thin'),
                bottom=Side(color='000000',border_style='thin'),
                diagonal=Side(color='000000',border_style='thin'),
            )
        )

        work_sheet.cell(row=row_num,column=6).value = '%.2f'%(all_pass_case_num/all_case_num*100)+'%' # "=C"+str(row_num)+"/B"+str(row_num)+""
        work_sheet.cell(row=row_num,column=6).style = Style(
            font=Font(
                name=u'微软雅黑',
                size=10
            ),
            border=Border(
                left=Side(color='000000',border_style='thin'),
                right=Side(color='000000',border_style='thin'),
                top=Side(color='000000',border_style='thin'),
                bottom=Side(color='000000',border_style='thin'),
                diagonal=Side(color='000000',border_style='thin'),
            )
        )



        for row in range(4,interface_num+4):
            work_sheet.cell(row=row,column=1).style = Style(
                font=Font(
                    name=u'微软雅黑',
                    size=10
                ),
                fill=PatternFill(
                    start_color='CCFFCC',
                    end_color='CCFFCC',
                    fill_type='solid'
                ),
                border=Border(
                    left=Side(color='000000',border_style='thin'),
                    right=Side(color='000000',border_style='thin'),
                    top=Side(color='000000',border_style='thin'),
                    bottom=Side(color='000000',border_style='thin'),
                    diagonal=Side(color='000000',border_style='thin'),
                )
            )
            work_sheet.cell(row=row,column=2).style = Style(
                font=Font(
                    name=u'微软雅黑',
                    size=10
                ),
                border=Border(
                    left=Side(color='000000',border_style='thin'),
                    right=Side(color='000000',border_style='thin'),
                    top=Side(color='000000',border_style='thin'),
                    bottom=Side(color='000000',border_style='thin'),
                    diagonal=Side(color='000000',border_style='thin'),
                )
            )
            work_sheet.cell(row=row,column=3).style = Style(
                font=Font(
                    name=u'微软雅黑',
                    size=10
                ),
                fill=PatternFill(
                    start_color='92D050',
                    end_color='92D050',
                    fill_type='solid'
                ),
                border=Border(
                    left=Side(color='000000',border_style='thin'),
                    right=Side(color='000000',border_style='thin'),
                    top=Side(color='000000',border_style='thin'),
                    bottom=Side(color='000000',border_style='thin'),
                    diagonal=Side(color='000000',border_style='thin'),
                )
            )
            work_sheet.cell(row=row,column=4).style = Style(
                font=Font(
                    name=u'微软雅黑',
                    size=10
                ),
                fill=PatternFill(
                    start_color='FF0000',
                    end_color='FF0000',
                    fill_type='solid'
                ),
                border=Border(
                    left=Side(color='000000',border_style='thin'),
                    right=Side(color='000000',border_style='thin'),
                    top=Side(color='000000',border_style='thin'),
                    bottom=Side(color='000000',border_style='thin'),
                    diagonal=Side(color='000000',border_style='thin'),
                )
            )
            work_sheet.cell(row=row,column=5).style = Style(
                font=Font(
                    name=u'微软雅黑',
                    size=10
                ),
                border=Border(
                    left=Side(color='000000',border_style='thin'),
                    right=Side(color='000000',border_style='thin'),
                    top=Side(color='000000',border_style='thin'),
                    bottom=Side(color='000000',border_style='thin'),
                    diagonal=Side(color='000000',border_style='thin'),
                )
            )
            work_sheet.cell(row=row,column=6).style = Style(
                font=Font(
                    name=u'微软雅黑',
                    size=10
                ),
                border=Border(
                    left=Side(color='000000',border_style='thin'),
                    right=Side(color='000000',border_style='thin'),
                    top=Side(color='000000',border_style='thin'),
                    bottom=Side(color='000000',border_style='thin'),
                    diagonal=Side(color='000000',border_style='thin'),
                )
            )


    # 读取excel
    def open_excel(self):
        # 所有sheet name
        names = self.work_book.get_sheet_names()
        for name in names:
            print name



class Excel_2003_Engine(object):
    def __init__(self,__name=None):
        # define class variable
        if not __name == None:
            self.xls_name = __name
            self.xlrd_object = None
            self.isopenfailed = True

    def open(self):
        try:
            self.xlrd_object = xlrd.open_workbook(self.xls_name)
            self.isopenfailed = False
            pass
        except:
            self.isopenfailed = True
            self.xlrd_object = None
            print "open %s failed \n"%self.xlrd_object
            pass
        finally:
            '''
            do nothing
            '''
            pass

        '''
        return a list
        '''
        return [self.isopenfailed,self.xlrd_object]

    def dump_sheet(self):
        if self.isopenfailed == False:
            for name in self.xlrd_object.sheet_names():
                table = self.xlrd_object.sheet_by_name(name)
                print "sheet %s rownums=%d colnums=%d"%(name,table.nrows,table.ncols)

    def dump_cell(self,sheet_index,cell_row,cell_col):
        try:
            table = self.xlrd_object.sheet_by_index(0)
            value = table.cell(cell_row,cell_col).value
            print "value=%s"%value
            pass
        except:
            pass

    def modify_cell(self,sheet_index,cell_row,cell_col,__value):
        try:
            table = self.xlrd_object.sheet_by_index(0)
            value = table.cell(cell_row,cell_col).value
            print "value=%s"%value
            table.put_cell(cell_row,cell_col,1,__value,0)
            value = table.cell(cell_row,cell_col).value
            print "value=%s"%value
            pass
        except:
            print "error"
        pass

    # 新建一个excel文档
    def create_excel(self):
        # excel对象
        workbook = xlwt.Workbook(encoding='utf-8')
        work_sheet = workbook.add_sheet('mysheet')
        work_sheet.write(0, 0, label = 'Row 0, Column 0 Value')
        workbook.save('../Download/Excel_Workbook.xls')

if __name__ == '__main__':
    t_2007 = Excel_2007_Engine('Test')
    t_2007.create_excel('Test')